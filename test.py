#Question 1
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

# Counter for matching records
count = 0

# Loop through rows and check conditions
for row in range(2, max_row + 1):
    address = ws['D' + str(row)].value
    quantity = ws['L' + str(row)].value

    if (
        isinstance(address, str) and address.startswith("Ain")
        and isinstance(quantity, (int, float)) and quantity < 40
    ):
        count += 1

print("Number of records with address starting 'Ain' and Skaits < 40:", count)


#Question 2
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

# Initialize counter
high_priority_2015_count = 0

# Loop through rows
for row in range(2, max_row + 1):
    priority = ws['H' + str(row)].value
    delivery_date = ws['J' + str(row)].value

    if (
        isinstance(priority, str) and priority.strip().lower() == "high"
        and delivery_date is not None
    ):
        # Ensure delivery_date is a date object or parse if it's string
        if hasattr(delivery_date, 'year') and delivery_date.year == 2015:
            high_priority_2015_count += 1
        elif isinstance(delivery_date, str) and delivery_date.strip().endswith("2015"):
            high_priority_2015_count += 1

print("Number of 'High' priority entries from 2015:", high_priority_2015_count)

#Question 3
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('data.xlsx') 
ws = wb['Lapa_0']
max_row = ws.max_row

# Count how many times "Adulienas iela" appears AND city is Valmiera or Saulkrasti
count = 0
for row in range(2, max_row + 1):
    address = ws['D' + str(row)].value
    city = ws['E' + str(row)].value

    if (
        isinstance(address, str) and "Adulienas iela" in address
        and isinstance(city, str) and city.strip() in ["Valmiera", "Saulkrasti"]
    ):
        count += 1

print("Number of 'Adulienas iela' entries in Valmiera or Saulkrasti:", count)

#Question 4
from openpyxl import load_workbook
import math

# Load the Excel file
wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

# Gather prices for products with "LaserJet" in the name
laserjet_prices = []
for row in range(2, max_row + 1):
    product = ws['I' + str(row)].value
    price = ws['K' + str(row)].value

    if isinstance(product, str) and "LaserJet" in product:
        if isinstance(price, (int, float)):
            laserjet_prices.append(price)

# Calculate average and round down
if laserjet_prices:
    average_price = sum(laserjet_prices) / len(laserjet_prices)
    average_price_rounded = math.floor(average_price)
    print("Average total of Produkts containing Laserjet:",average_price_rounded)  
else:
    print("0")

#Question 5
from openpyxl import load_workbook

# Load workbook and select sheet
wb = load_workbook('data.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

# Initialize total sum
total_sum = 0

# Loop through rows
for row in range(2, max_row + 1):
    client_type = ws['F' + str(row)].value
    quantity = ws['L' + str(row)].value
    total_raw = ws['N' + str(row)].value

    if (
        isinstance(client_type, str) and client_type.strip() == "Korporatīvais"
        and isinstance(quantity, (int, float)) and 40 <= quantity <= 50
    ):
        # Clean total value
        if isinstance(total_raw, str):
            cleaned = ''.join(ch for ch in total_raw if ch.isdigit() or ch == '.' or ch == ',')
            cleaned = cleaned.replace(',', '') 
            try:
                total = float(cleaned)
            except ValueError:
                continue 
        elif isinstance(total_raw, (int, float)):
            total = total_raw
        else:
            continue

        total_sum += total

print("Total Kopā for 'Korporatīvais' clients with Skaits between 40–50:", int(total_sum))
